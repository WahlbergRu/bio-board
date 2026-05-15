"""Excel import/export for Gantt planner tasks."""

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook, load_workbook


# Column name aliases (Russian primary, English fallback)
_COL_MAP: dict[str, list[str]] = {
    "name": ["задача", "task", "название", "name"],
    "description": ["описание", "description", "desc", "описание задачи"],
    "assignee": ["исполнитель", "assignee", "responsible", "ответственный"],
    "duration": ["длительность", "duration", "дней", "days"],
    "predecessors": ["предшественники", "predecessors", "зависимости", "dependencies"],
}


def _resolve_columns(headers: list[str]) -> dict[str, int]:
    """Map canonical field names to column indices."""
    mapping: dict[str, int] = {}
    lower_headers = [h.strip().lower() for h in headers]
    for field, aliases in _COL_MAP.items():
        for alias in aliases:
            if alias in lower_headers:
                mapping[field] = lower_headers.index(alias)
                break
    return mapping


def _safe_str(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _calculate_duration(start_date: str, end_date: str) -> int:
    """Days between two ISO date strings (inclusive of start)."""
    try:
        s = datetime.fromisoformat(start_date).date()
        e = datetime.fromisoformat(end_date).date()
        return max((e - s).days, 0)
    except (ValueError, TypeError):
        return 0


def parse_excel(file_bytes: bytes) -> list[dict]:
    """Parse xlsx bytes into list of task dicts.

    Returns dicts with keys: name, description, assignee, duration, predecessors.
    """
    try:
        wb = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        return []
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    headers = [_safe_str(h) for h in rows[0]]
    cols = _resolve_columns(headers)
    result: list[dict] = []

    for row in rows[1:]:
        if not row or all(c is None for c in row):
            continue

        def _get(field: str) -> str:
            idx = cols.get(field)
            if idx is None or idx >= len(row):
                return ""
            return _safe_str(row[idx])

        pred_raw = _get("predecessors")
        predecessors = [p.strip() for p in pred_raw.split(";") if p.strip()] if pred_raw else []

        dur_raw = _get("duration")
        try:
            duration = int(float(dur_raw)) if dur_raw else 0
        except ValueError:
            duration = 0

        result.append({
            "name": _get("name"),
            "description": _get("description"),
            "assignee": _get("assignee"),
            "duration": duration,
            "predecessors": predecessors,
        })

    return result


def export_excel(tasks: list[dict]) -> bytes:
    """Export task dicts to xlsx bytes.

    Expected task keys: name, description, assignee, start_date, end_date,
    predecessors (list of str). Duration is calculated from dates.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "План"

    columns = ["задача", "описание", "исполнитель", "длительность", "предшественники"]
    ws.append(columns)

    for t in tasks:
        start = t.get("start_date", "")
        end = t.get("end_date", "")
        duration = _calculate_duration(start, end) if start and end else t.get("duration", 0)
        preds = ";".join(t.get("predecessors", []))

        ws.append([
            t.get("name", ""),
            t.get("description", ""),
            t.get("assignee", ""),
            duration,
            preds,
        ])

    buf = BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()
