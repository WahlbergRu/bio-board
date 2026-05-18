"""Route integration tests using FastAPI TestClient."""

import pytest
from fastapi.testclient import TestClient
from unittest.mock import patch, AsyncMock
import io

from app.main import app, rate_limits


@pytest.fixture
def client():
    """Create test client with clean state."""
    app.state.store.tasks.clear()
    app.state.store.save()
    rate_limits.clear()
    yield TestClient(app)
    app.state.store.tasks.clear()
    app.state.store.save()


# ── Health ────────────────────────────────────────────────────────

def test_health(client):
    res = client.get("/health")
    assert res.status_code == 200
    assert res.json()["status"] == "ok"


# ── Tasks CRUD ────────────────────────────────────────────────────

def test_list_tasks_empty(client):
    res = client.get("/api/tasks/")
    assert res.status_code == 200
    assert res.json() == []


def test_create_task(client):
    res = client.post("/api/tasks/", json={
        "name": "Test Task",
        "start_date": "2026-01-01",
        "end_date": "2026-01-05",
        "assignee": "Alice",
    })
    assert res.status_code == 201
    data = res.json()
    assert data["name"] == "Test Task"
    assert data["assignee"] == "Alice"
    assert data["id"] is not None


def test_get_task(client):
    create_res = client.post("/api/tasks/", json={
        "name": "Get Me", "start_date": "2026-01-01", "end_date": "2026-01-05",
    })
    task_id = create_res.json()["id"]
    res = client.get(f"/api/tasks/{task_id}")
    assert res.status_code == 200
    assert res.json()["name"] == "Get Me"


def test_get_task_not_found(client):
    res = client.get("/api/tasks/nonexistent")
    assert res.status_code == 404


def test_update_task(client):
    create_res = client.post("/api/tasks/", json={
        "name": "Old Name", "start_date": "2026-01-01", "end_date": "2026-01-05",
    })
    task_id = create_res.json()["id"]
    res = client.put(f"/api/tasks/{task_id}", json={
        "id": task_id, "name": "New Name",
    })
    assert res.status_code == 200
    assert res.json()["name"] == "New Name"


def test_delete_task(client):
    create_res = client.post("/api/tasks/", json={
        "name": "Delete Me", "start_date": "2026-01-01", "end_date": "2026-01-05",
    })
    task_id = create_res.json()["id"]
    res = client.delete(f"/api/tasks/{task_id}")
    assert res.status_code == 204
    get_res = client.get(f"/api/tasks/{task_id}")
    assert get_res.status_code == 404


def test_create_task_xss_sanitized(client):
    res = client.post("/api/tasks/", json={
        "name": "<script>alert('xss')</script>",
        "start_date": "2026-01-01",
        "end_date": "2026-01-05",
    })
    assert res.status_code == 201
    assert "<script>" not in res.json()["name"]


# ── Plan ──────────────────────────────────────────────────────────

def test_seed_plan(client):
    res = client.post("/api/plan/seed")
    assert res.status_code == 200
    tasks = client.get("/api/tasks/").json()
    assert len(tasks) >= 3


def test_reset_plan(client):
    client.post("/api/plan/seed")
    res = client.request("DELETE", "/api/plan/reset")
    assert res.status_code == 200
    tasks = client.get("/api/tasks/").json()
    assert len(tasks) == 0


def test_get_plan(client):
    client.post("/api/plan/seed")
    res = client.get("/api/plan/")
    assert res.status_code == 200
    data = res.json()
    assert "tasks" in data or isinstance(data, list)


# ── Excel Export ──────────────────────────────────────────────────

def test_export_excel(client):
    client.post("/api/tasks/", json={
        "name": "Export Test", "start_date": "2026-01-01", "end_date": "2026-01-05",
    })
    res = client.get("/api/excel/export")
    assert res.status_code == 200
    assert res.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument"
    )


# ── Excel Upload ──────────────────────────────────────────────────

def test_upload_excel(client, sample_tasks_xlsx):
    with open(sample_tasks_xlsx, "rb") as f:
        res = client.post("/api/excel/upload", files={"file": ("test.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert res.status_code == 200
    data = res.json()
    assert "imported" in data
    assert data["imported"] >= 1


@pytest.fixture
def sample_tasks_xlsx(tmp_path):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["задача", "описание", "исполнитель", "длительность", "предшественники"])
    ws.append(["Task A", "Description A", "Alice", 3, ""])
    ws.append(["Task B", "Description B", "Bob", 5, "1"])
    fpath = str(tmp_path / "test.xlsx")
    wb.save(fpath)
    return fpath


@pytest.fixture
def sample_tasks_with_deps_xlsx(tmp_path):
    """Excel with predecessor names (not IDs) for roundtrip testing."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["задача", "описание", "исполнитель", "длительность", "предшественники"])
    ws.append(["Анализ", "Сбор требований", "Аналитик", 5, ""])
    ws.append(["Дизайн", "Проектирование", "Архитектор", 7, "Анализ"])
    ws.append(["Разработка", "Кодирование", "Разработчик", 10, "Дизайн"])
    ws.append(["Тест", "Тестирование", "QA", 3, "Разработка"])
    fpath = str(tmp_path / "deps_test.xlsx")
    wb.save(fpath)
    return fpath


# ── Excel Roundtrip ───────────────────────────────────────────────

def _extract_tasks_from_xlsx(xlsx_bytes):
    """Parse xlsx bytes into list of (name, predecessors) tuples."""
    from openpyxl import load_workbook
    from io import BytesIO
    wb = load_workbook(filename=BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    result = []
    for row in rows[1:]:
        if not row or all(c is None for c in row):
            continue
        name = str(row[0]).strip() if row[0] else ""
        preds = str(row[4]).strip() if row[4] else ""
        result.append((name, preds))
    return result


def test_excel_upload_export_roundtrip(client, sample_tasks_with_deps_xlsx):
    """Upload Excel -> Export -> Verify predecessors preserved by name."""
    # Upload
    with open(sample_tasks_with_deps_xlsx, "rb") as f:
        res = client.post("/api/excel/upload", files={"file": ("test.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert res.status_code == 200
    assert res.json()["imported"] == 4

    # Export
    res = client.get("/api/excel/export")
    assert res.status_code == 200
    tasks1 = _extract_tasks_from_xlsx(res.content)
    assert len(tasks1) == 4

    # Verify predecessors are task names, not empty
    names = [t[0] for t in tasks1]
    assert "Анализ" in names
    assert "Дизайн" in names
    assert "Разработка" in names
    assert "Тест" in names

    # Find each task and check its predecessor column
    task_map = {name: preds for name, preds in tasks1}
    assert task_map["Анализ"] == ""
    assert task_map["Дизайн"] == "Анализ"
    assert task_map["Разработка"] == "Дизайн"
    assert task_map["Тест"] == "Разработка"

    # Reset and re-import the exported file
    client.request("DELETE", "/api/plan/reset")

    import tempfile
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        tf.write(res.content)
        tf.flush()
        tf_name = tf.name

    with open(tf_name, "rb") as f:
        res2 = client.post("/api/excel/upload", files={"file": ("roundtrip.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert res2.status_code == 200
    assert res2.json()["imported"] == 4

    # Export again
    res3 = client.get("/api/excel/export")
    assert res3.status_code == 200
    tasks2 = _extract_tasks_from_xlsx(res3.content)
    assert len(tasks2) == 4

    # Verify structure is identical
    task_map2 = {name: preds for name, preds in tasks2}
    assert task_map2["Анализ"] == ""
    assert task_map2["Дизайн"] == "Анализ"
    assert task_map2["Разработка"] == "Дизайн"
    assert task_map2["Тест"] == "Разработка"

    import os
    os.unlink(tf_name)


def test_excel_complex_deps_roundtrip(client):
    """Upload complex Excel with multiple deps -> Export -> Re-import -> Export -> Compare."""
    from openpyxl import Workbook
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = "План"
    ws.append(["задача", "описание", "исполнитель", "длительность", "предшественники"])
    ws.append(["Старт", "Начало", "PM", 1, ""])
    ws.append(["ДБ", "База данных", "Backend", 4, "Старт"])
    ws.append(["API", "API дизайн", "Backend", 4, "Старт"])
    ws.append(["UI", "Интерфейс", "Frontend", 6, "Старт"])
    ws.append(["Auth", "Авторизация", "Backend", 5, "ДБ;API"])
    ws.append(["Core", "Ядро", "Backend", 10, "ДБ;API"])
    ws.append(["Front", "Фронтенд", "Frontend", 8, "UI"])
    ws.append(["Integ", "Интеграция", "Fullstack", 5, "Auth;Core;Front"])
    ws.append(["Тест", "Тесты", "QA", 4, "Integ"])
    ws.append(["Релиз", "Деплой", "DevOps", 1, "Тест"])
    buf = BytesIO()
    wb.save(buf)
    wb.close()

    # Upload
    res = client.post("/api/excel/upload", files={"file": ("complex.xlsx", buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert res.status_code == 200
    assert res.json()["imported"] == 10

    # Export round 1
    res = client.get("/api/excel/export")
    assert res.status_code == 200
    tasks1 = _extract_tasks_from_xlsx(res.content)
    assert len(tasks1) == 10

    # Reset and re-import
    client.request("DELETE", "/api/plan/reset")

    import tempfile
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        tf.write(res.content)
        tf.flush()
        tf_name = tf.name

    with open(tf_name, "rb") as f:
        res2 = client.post("/api/excel/upload", files={"file": ("round2.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert res2.status_code == 200
    assert res2.json()["imported"] == 10

    # Export round 2
    res3 = client.get("/api/excel/export")
    assert res3.status_code == 200
    tasks2 = _extract_tasks_from_xlsx(res3.content)
    assert len(tasks2) == 10

    # Compare full structure
    map1 = {name: preds for name, preds in tasks1}
    map2 = {name: preds for name, preds in tasks2}
    assert map1 == map2

    # Verify specific complex deps
    assert map2["Auth"] == "ДБ;API"
    assert map2["Core"] == "ДБ;API"
    assert map2["Integ"] == "Auth;Core;Front"
    assert map2["Тест"] == "Integ"
    assert map2["Релиз"] == "Тест"

    import os
    os.unlink(tf_name)


# ── Chat ──────────────────────────────────────────────────────────

def test_chat_endpoint(client):
    """Chat endpoint accepts requests and returns SSE stream."""
    res = client.post("/api/chat/", json={
        "message": "Add a task",
        "history": [],
    })
    # SSE endpoint — should return 200 with streaming content
    assert res.status_code == 200
    assert "text/event-stream" in res.headers.get("content-type", "")


# ── Auth ──────────────────────────────────────────────────────────

def test_login_success(client):
    res = client.post("/api/auth/login", json={
        "username": "admin",
        "password": "admin",
    })
    assert res.status_code == 200
    data = res.json()
    assert "access_token" in data


def test_login_failure(client):
    res = client.post("/api/auth/login", json={
        "username": "admin",
        "password": "wrong",
    })
    assert res.status_code == 401


def test_get_me_with_token(client):
    login_res = client.post("/api/auth/login", json={
        "username": "admin", "password": "admin",
    })
    token = login_res.json()["access_token"]
    res = client.get("/api/auth/me", headers={"Authorization": f"Bearer {token}"})
    assert res.status_code == 200
    assert res.json()["user"] == "admin"


def test_get_me_no_token(client):
    res = client.get("/api/auth/me")
    assert res.status_code == 401
